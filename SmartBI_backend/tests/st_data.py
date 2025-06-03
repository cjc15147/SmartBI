import streamlit as st
import pandas as pd
import folium
from haversine import haversine
from geopy.distance import geodesic
import streamlit.components.v1 as components
from openpyxl import load_workbook
import plotly.express as px
from openpyxl.styles import PatternFill
import numpy as np

# 设置存储全部范围内存量基站的全局变量
global_nearby_df = pd.DataFrame()
# 通过经纬度计算距离
def haversine_distance(row, new_lat, new_lon):
    return haversine((row['纬度'], row['经度']), (new_lat, new_lon))

# 在循环之前，先为 df_cleaned 计算所有点到新基站的距离
def preprocess_distances(df_cleaned, new_stations_df):
    # 创建一个新的 DataFrame 来存储距离
    # distances_df = pd.DataFrame(index=df_cleaned.index)
    distances_df = df_cleaned.copy()

    for index, new_station in new_stations_df.iterrows():
        new_lat = new_station['纬度']
        new_lon = new_station['经度']
        distances_df[f'距离_{index+2}'] = df_cleaned.apply(
            lambda row: haversine_distance(row, new_lat, new_lon),
            axis=1
        )
    return distances_df

# 定义一个函数来创建地图
def create_map(center, distances_df,df1,df2,radius_km):
    map_ = folium.Map(location=center, zoom_start=12)

    # 将筛选后的存量机房添加到地图上（蓝色点）
    for index, row in df2.iterrows():
        # 使用预先计算的距离
        nearby_distances = distances_df[f'距离_{row.name + 2}']
        # nearby_distances.to_excel('车市.xlsx', index=False)
        # 筛选范围内存量机房的索引值
        filtered_indices = [i for i, value in enumerate(nearby_distances) if value <= radius_km]
        nearby_df = df1.iloc[filtered_indices]
        # 删除大于阈值的行
        nearby_df = nearby_df[nearby_df['合同年租金'] < 30000]
        for _, row in nearby_df.iterrows():
           if pd.notnull(row['纬度']) and pd.notnull(row['经度']):
              folium.Marker(
                location=[row['纬度'], row['经度']],
                popup=f"合同年租金: {row['合同年租金']}",
                icon=folium.Icon(color='blue')
              ).add_to(map_)

    # 将新增机房添加到地图上（红色点）
    for _, row in df2.iterrows():
        if pd.notnull(row['纬度']) and pd.notnull(row['经度']):
            folium.Marker(
                location=[row['纬度'], row['经度']],
                popup=f"合同年租金: {row['合同年租金']}",
                icon=folium.Icon(color='red')
            ).add_to(map_)

    return map_

# 创建散点图
def create_scatter_plot(distances_df,df1, df2, radius):

    # 从 df2 中选择特定列，并添加“类型”列
    df2_selected = df2[['合同年租金', '经度', '纬度']].copy()
    df2_selected['类型'] = '新增机房'

    # 从 df_filtered 中选择特定列，并添加“类型”列
    df1_selected = df1[['合同年租金', '经度', '纬度']].copy()
    df1_selected['类型'] = '存量机房'

    df_combined = pd.DataFrame(columns=['合同年租金', '经度', '纬度', '类型'])
    # 合并两类机房的散点数据后的DataFrame
    for index, row in df2.iterrows():
        # 使用预先计算的距离合集，将该基站的距离信息分离出来
        nearby_distances = distances_df[f'距离_{row.name + 2}']
        # nearby_distances.to_excel('车市.xlsx', index=False)
        # 筛选范围内存量机房的索引值
        filtered_indices = [i for i, value in enumerate(nearby_distances) if value <= radius]
        nearby_df = df1_selected.iloc[filtered_indices]
        # 删除大于阈值的行
        nearby_df = nearby_df[nearby_df['合同年租金'] < 30000]
        df_combined = pd.concat([df_combined, nearby_df], ignore_index=True)

    df_combined = pd.concat([df_combined, df2_selected], ignore_index=True)
    fig = px.scatter(df_combined,
                     x='经度',
                     y='纬度',
                     color='类型',
                     size='合同年租金',
                     color_discrete_map={'新增机房': 'red', '存量机房': 'blue'},
                     hover_data=['合同年租金'],
                     title='租金分布散点图')

    fig.update_layout(
        title_x=0,
        title_y=0.95,
        title_font_size=20,
        showlegend=False
    )

    return fig

# 定义一个函数来处理每个新基站的数据（不需要重新计算距离）
def process_new_station(row, distances_df,df1, radius_km):
    new_lat = row['纬度']
    new_lon = row['经度']
    new_rental = row['合同年租金']
    global global_nearby_df
    # 使用预先计算的距离

    nearby_distances = distances_df[f'距离_{row.name+2}']
    # nearby_distances.to_excel('车市.xlsx', index=False)
    # 筛选范围内存量机房的索引值
    filtered_indices = [i for i, value in enumerate(nearby_distances) if value <= radius_km]
    nearby_df = df1.iloc[filtered_indices]
    # 删除大于阈值的行
    nearby_df = nearby_df[nearby_df['合同年租金'] < 30000]
    # 筛选nearby_distances
    filtered_elements = nearby_distances.iloc[filtered_indices]
    # nearby_df.loc[:, f'距离_{row.name + 2}'] = nearby_distances.iloc[filtered_indices]
    # 给nearby_df后面加一列距离
    nearby_df['距离'] = np.nan
    nearby_df['距离'] = filtered_elements

    # 计算平均价格、最大值、最小值
    avg_rental = nearby_df['合同年租金'].mean()
    min_rental = nearby_df['合同年租金'].min()
    max_rental = nearby_df['合同年租金'].max()
    # nearest_rental = nearby_df['合同年租金'].iloc[0] if not nearby_df.empty else None
    #
    # 筛选出非零的'距离'值
    non_zero_distances = nearby_df[nearby_df['距离'] != 0]

    # 如果非零的'距离'存在，找到最小的那个，并获取对应的'合同年租金'
    if not non_zero_distances.empty:
        min_distance_idx = non_zero_distances['距离'].idxmin()  # 获取最小'距离'值的索引
        nearest_rental = non_zero_distances.at[min_distance_idx, '合同年租金']  # 使用索引获取'合同年租金'
    else:
        nearest_rental = None  # 如果没有非零的'距离'，则设置为None
    # nearest_rental = nearby_df.at[nearby_df['距离'].idxmin(), '合同年租金']

    # 给全局变量赋值
    # 如果global_df为空，则直接赋值
    if global_nearby_df.empty:
        global_nearby_df = nearby_df.copy()
    else:
        # 否则，使用pd.concat追加
        global_nearby_df = pd.concat([global_nearby_df, nearby_df], ignore_index=True)
        # 返回结果
    return {
        '新基站经度': new_lon,
        '新基站纬度': new_lat,
        '新基站合同年租金': new_rental,
        '周围存量机房平均合同年租金': avg_rental,
        '周围存量机房最小合同年租金': min_rental,
        '周围存量机房最大合同年租金': max_rental,
        '最近存量机房合同年租金': nearest_rental
    }


def main():
    st.title('新增机房报价自动稽核系统')

    with st.sidebar:
        st.header('上传文件')
        uploaded_file_1 = st.file_uploader("上传报账点信息管理Excel文档", type=['xlsx'])
        uploaded_file_2 = st.file_uploader("上传新机房报价汇总Excel表", type=['xlsx'])
        radius = st.number_input("请输入给定的半径（公里）：", min_value=1, max_value=100, value=10, step=1)
        audit_button = st.button("开始自动稽核")

    if uploaded_file_1 is not None and uploaded_file_2 is not None and audit_button:


        df1 = pd.read_excel(uploaded_file_1)
        df1 = df1.dropna(subset=['机房面积', '经度', '纬度', '合同年租金'])
        df2 = pd.read_excel(uploaded_file_2)
        df2 = df2.dropna(subset=['机房面积', '经度', '纬度', '合同年租金'])
        # 计算距离,将每一个新增机房距离所有存量机房的距离依次续在df1的后面列得到distances_df
        distances_df = preprocess_distances(df1, df2)



        xian_center = [34.341575, 108.93977]

        map_ = create_map(xian_center, distances_df,df1, df2,radius)

        st.subheader("地图展示")
        map_html = map_._repr_html_()
        components.html(map_html, height=600)

        scatter_fig = create_scatter_plot(distances_df,df1, df2, radius)

        # st.subheader("租金分布散点图")
        st.plotly_chart(scatter_fig, use_container_width=True)


        #新增机房excel比价结果展示


        # 复制new_stations_quotes到results_df
        # results_df = df2.copy()
        columns_to_select = ['机房面积', '经度', '纬度', '合同年租金']
        results_df = df2[columns_to_select].copy()

        # 添加新列，初始值为NaN
        new_columns = ['周围存量机房平均合同年租金', '周围存量机房最小合同年租金', '周围存量机房最大合同年租金',
                       '最近存量机房合同年租金']
        for column in new_columns:
            results_df[column] = np.nan

        for index, row in results_df.iterrows():
            # 这里假设process_new_station可以直接处理row
            # 注意：在实际应用中，可能需要将row转换为process_new_station所需的格式
            rental_data = process_new_station(row, distances_df, df1, radius)  # 这里可能需要调整以适应process_new_station的实际输入

            # 更新DataFrame中的列
            results_df.at[index, '周围存量机房平均合同年租金'] = rental_data['周围存量机房平均合同年租金']
            results_df.at[index, '周围存量机房最小合同年租金'] = rental_data['周围存量机房最小合同年租金']
            results_df.at[index, '周围存量机房最大合同年租金'] = rental_data['周围存量机房最大合同年租金']
            results_df.at[index, '最近存量机房合同年租金'] = rental_data['最近存量机房合同年租金']
        results_df['新机房租金与周边平均租金对比结果'] = ['<=' if a <= b else '>' for a, b in
                                                          zip(results_df['合同年租金'],
                                                              results_df['周围存量机房平均合同年租金'])]
        results_df['新机房租金与最近存量机房租金对比结果'] = ['<=' if a <= b else '>' for a, b in
                                                              zip(results_df['合同年租金'],
                                                                  results_df['最近存量机房合同年租金'])]

        results_df.to_excel('新机房周边的机房数据对比.xlsx', index=False)
        # global_nearby_df.to_excel('teste.xlsx', index=False)
        # 给新机房周边的机房数据对比.xlsx打标识
        wb = load_workbook('新机房周边的机房数据对比.xlsx')
        ws = wb.active

        # 遍历除了标题行之外的所有行
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # 获取'对比结果'列的单元格
            cell_result_1 = row[results_df.columns.get_loc('新机房租金与周边平均租金对比结果')]
            cell_result_2 = row[results_df.columns.get_loc('新机房租金与最近存量机房租金对比结果')]

            # 根据'对比结果'的值来设置单元格的背景色
            if cell_result_1.value == '<=':
                cell_result_1.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                cell_result_1.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

            if cell_result_2.value == '<=':
                cell_result_2.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                cell_result_2.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

            # 保存修改后的Excel文件
        wb.save('新机房周边的机房数据对比.xlsx')
        df = pd.read_excel('新机房周边的机房数据对比.xlsx')

        # 显示前几行数据
        st.write("新机房报价对比结果汇总:")
        st.dataframe(df.head())



if __name__ == '__main__':
    main()